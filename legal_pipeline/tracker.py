"""tracker.py — writes the migration_tracker.xlsx report."""
from __future__ import annotations
from pathlib import Path
from typing import List, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

_PREFERRED_COLUMNS = [
    "idx", "file_name", "doc_type", "confidence", "needs_review",
    "status", "is_duplicate", "duplicate_of", "dup_method", "similarity",
    "entity_name", "contract_date", "description",
    "page_count", "size_kb", "extraction_method", "is_scanned",
    "target_folder", "consolidated_path",
    "path", "file_sha256", "text_sha256", "score", "matched_terms",
]

_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_REVIEW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_REMOVED_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")


def _ordered_columns(rows: List[dict]) -> List[str]:
    seen = set()
    cols = [c for c in _PREFERRED_COLUMNS if any(c in r for r in rows)]
    seen.update(cols)
    for r in rows:
        for k in r.keys():
            if k not in seen:
                cols.append(k)
                seen.add(k)
    return cols


def _write_tracker_sheet(ws, rows: List[dict]):
    cols = _ordered_columns(rows)
    ws.append(cols)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")

    for r in rows:
        values = []
        for c in cols:
            v = r.get(c, "")
            if isinstance(v, (list, tuple)):
                v = ", ".join(str(x) for x in v)
            values.append(v)
        ws.append(values)

    for row_idx, r in enumerate(rows, start=2):
        fill = None
        if r.get("status") == "removed":
            fill = _REMOVED_FILL
        elif r.get("needs_review"):
            fill = _REVIEW_FILL
        if fill:
            for col_idx in range(1, len(cols) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    for col_idx, c in enumerate(cols, start=1):
        max_len = max([len(c)] + [len(str(r.get(c, ""))) for r in rows[:500]])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 50)

    ws.freeze_panes = "A2"

    if rows:
        end_col = get_column_letter(len(cols))
        table_ref = f"A1:{end_col}{len(rows) + 1}"
        table = Table(displayName="TrackerTable", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False
        )
        ws.add_table(table)


def _write_summary_sheet(ws, rows: List[dict], ocr_note: str,
                          consolidated_dir: Optional[Path]):
    total = len(rows)
    retained = sum(1 for r in rows if r.get("status") == "retained")
    removed = total - retained
    needs_review = sum(1 for r in rows if r.get("needs_review"))
    scanned = sum(1 for r in rows if r.get("is_scanned"))
    failed = sum(1 for r in rows if r.get("extraction_method") == "failed")

    ws.append(["Legal Document Migration — Summary"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    pairs = [
        ("Total documents scanned", total),
        ("Retained (to migrate)", retained),
        ("Removed as duplicates", removed),
        ("Flagged for manual review", needs_review),
        ("Scanned pages (OCR candidates)", scanned),
        ("Files that failed extraction entirely", failed),
        ("OCR engine status", ocr_note),
        ("Consolidated folder (ALL retained files)", str(consolidated_dir) if consolidated_dir else ""),
    ]
    for label, value in pairs:
        ws.append([label, value])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 60


def build_tracker(rows: List[dict], tracker_path: Path,
                   ocr_note: str = "", consolidated_dir: Optional[Path] = None) -> Path:
    tracker_path = Path(tracker_path)
    tracker_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"
    _write_summary_sheet(summary_ws, rows, ocr_note, consolidated_dir)

    tracker_ws = wb.create_sheet("Tracker")
    _write_tracker_sheet(tracker_ws, rows)

    wb.active = 0
    wb.save(str(tracker_path))
    return tracker_path
