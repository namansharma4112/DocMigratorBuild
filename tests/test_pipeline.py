"""End-to-end and unit tests for the legal document migration pipeline."""
from __future__ import annotations

import shutil
from pathlib import Path

import pytest
from openpyxl import load_workbook

from legal_pipeline.classify import classify_document
from legal_pipeline.config import Config, ClassificationThresholds, Dedup
from legal_pipeline.dedupe import DedupRecord, deduplicate
from legal_pipeline.metadata import extract_metadata
from legal_pipeline.pipeline import run

from .make_test_pdfs import main as build_standard_fixture_set, OUT as STANDARD_FIXTURE_DIR


@pytest.fixture(scope="module")
def fixture_pdfs() -> Path:
    build_standard_fixture_set()
    return STANDARD_FIXTURE_DIR


@pytest.fixture()
def output_dir(tmp_path) -> Path:
    d = tmp_path / "output"
    yield d
    if d.exists():
        shutil.rmtree(d, ignore_errors=True)


# --------------------------------------------------------------------------- #
# End-to-end pipeline tests
# --------------------------------------------------------------------------- #
def test_full_pipeline_runs_without_crashing_and_produces_expected_counts(fixture_pdfs, output_dir):
    cfg = Config()
    cfg.paths.source_dir = fixture_pdfs
    cfg.paths.output_dir = output_dir

    summary = run(cfg, copy_files=True, log=lambda *_: None)

    # 12 fixture PDFs: #02 is an exact byte-copy of #01, #03 is a native-text
    # near-dup of #01 (deterministic - no OCR involved), #08 is an exact
    # byte-copy of #07 (see make_test_pdfs.py's module docstring for why #08
    # is a copy rather than a re-rendered OCR near-duplicate) -> 3 removed,
    # 9 retained. This count is now deterministic on every platform.
    assert summary["total"] == 12
    assert summary["removed"] == 3
    assert summary["retained"] == 9

    tracker_path = Path(summary["tracker"])
    assert tracker_path.exists()
    consolidated = Path(summary["consolidated_dir"])
    assert consolidated.exists()
    assert sum(1 for _ in consolidated.glob("*.pdf")) == 9


def test_no_vendor_cap_fields_anywhere(fixture_pdfs, output_dir):
    cfg = Config()
    cfg.paths.source_dir = fixture_pdfs
    cfg.paths.output_dir = output_dir
    summary = run(cfg, copy_files=True, log=lambda *_: None)

    assert "vendor_cap" not in summary
    assert "within_cap" not in summary
    assert "headroom" not in summary
    assert not hasattr(cfg, "migration")

    wb = load_workbook(summary["tracker"])
    summary_ws = wb["Summary"]
    label_text = "\n".join(str(row[0].value) for row in summary_ws.iter_rows() if row[0].value)
    assert "vendor cap" not in label_text.lower()
    assert "within cap" not in label_text.lower()
    assert "headroom" not in label_text.lower()


def test_corrupt_and_unreadable_files_never_crash_the_run(fixture_pdfs, output_dir, tmp_path):
    import os

    src = tmp_path / "with_unreadable"
    src.mkdir()
    for p in fixture_pdfs.glob("*.pdf"):
        shutil.copy(p, src / p.name)
    bad = src / "unreadable_locked_file.pdf"
    bad.write_bytes(b"%PDF-1.4 pretend content")
    os.chmod(bad, 0o000)
    try:
        cfg = Config()
        cfg.paths.source_dir = src
        cfg.paths.output_dir = output_dir
        summary = run(cfg, copy_files=True, log=lambda *_: None)
        assert summary["total"] == 13
        assert summary["failed_extraction"] >= 1
    finally:
        os.chmod(bad, 0o644)


def test_extraction_cache_invalidates_on_content_change_despite_identical_size(output_dir, tmp_path):
    import time
    from legal_pipeline.pipeline import stage_extract

    src = tmp_path / "cache_src"
    src.mkdir()
    pdf_path = src / "test.pdf"

    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4

    def make_pdf_padded(text, target_size=None):
        c = canvas.Canvas(str(pdf_path), pagesize=A4)
        c.drawString(50, 800, text)
        c.save()
        if target_size:
            data = pdf_path.read_bytes()
            if len(data) < target_size:
                pad = b"%" + b"x" * (target_size - len(data) - 2) + b"\n"
                data = data[:-6] + pad + data[-6:]
                pdf_path.write_bytes(data)

    make_pdf_padded("ORIGINAL TEXT AAAA")
    size1 = pdf_path.stat().st_size

    cfg = Config()
    cfg.paths.output_dir = output_dir
    docs1 = stage_extract(cfg, [pdf_path])
    assert "ORIGINAL" in docs1[0].text

    time.sleep(1.1)
    make_pdf_padded("DIFFERENT TEXT BBBB", target_size=size1)
    size2 = pdf_path.stat().st_size
    assert size1 == size2

    docs2 = stage_extract(cfg, [pdf_path])
    assert "DIFFERENT" in docs2[0].text


def test_extraction_cache_still_hits_for_genuinely_unchanged_files(output_dir, tmp_path):
    from legal_pipeline.pipeline import stage_extract

    src = tmp_path / "cache_src2"
    src.mkdir()
    pdf_path = src / "test.pdf"
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    c = canvas.Canvas(str(pdf_path), pagesize=A4)
    c.drawString(50, 800, "STABLE TEXT")
    c.save()

    cfg = Config()
    cfg.paths.output_dir = output_dir
    docs1 = stage_extract(cfg, [pdf_path])
    docs2 = stage_extract(cfg, [pdf_path])
    assert docs1[0].text == docs2[0].text
    assert docs1[0].file_sha256 == docs2[0].file_sha256


def test_rerun_on_same_output_dir_does_not_duplicate_files(tmp_path):
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4

    src = tmp_path / "src"
    src.mkdir()
    for i, folder in enumerate(["f1", "f2", "f3"]):
        d = src / folder
        d.mkdir()
        c = canvas.Canvas(str(d / "contract.pdf"), pagesize=A4)
        c.drawString(50, 800, f"SERVICES AGREEMENT unique client {i} distinct content {i}{i}{i}")
        c.save()

    out = tmp_path / "out"
    cfg = Config()
    cfg.paths.source_dir = src
    cfg.paths.output_dir = out

    run(cfg, copy_files=True, log=lambda *_: None)
    first_run_files = sorted(f.name for f in (out / "Consolidated").glob("*.pdf"))
    assert len(first_run_files) == 3

    run(cfg, copy_files=True, log=lambda *_: None)
    second_run_files = sorted(f.name for f in (out / "Consolidated").glob("*.pdf"))
    assert second_run_files == first_run_files


def test_rerun_removes_stale_files_no_longer_present_in_source(tmp_path):
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4

    src = tmp_path / "src"
    src.mkdir()
    c = canvas.Canvas(str(src / "doc_a.pdf"), pagesize=A4)
    c.drawString(50, 800, "SERVICES AGREEMENT unique client Alpha Holdings LLC content one")
    c.save()
    c = canvas.Canvas(str(src / "doc_b.pdf"), pagesize=A4)
    c.drawString(50, 800, "SERVICES AGREEMENT unique client Beta Holdings LLC content two different")
    c.save()

    out = tmp_path / "out"
    cfg = Config()
    cfg.paths.source_dir = src
    cfg.paths.output_dir = out

    summary1 = run(cfg, copy_files=True, log=lambda *_: None)
    assert summary1["retained"] == 2

    (src / "doc_b.pdf").unlink()

    summary2 = run(cfg, copy_files=True, log=lambda *_: None)
    assert summary2["retained"] == 1

    consolidated_after = sorted(f.name for f in (out / "Consolidated").glob("*.pdf"))
    assert consolidated_after == ["doc_a.pdf"]


def test_different_client_boilerplate_contract_is_not_falsely_merged(fixture_pdfs, output_dir):
    cfg = Config()
    cfg.paths.source_dir = fixture_pdfs
    cfg.paths.output_dir = output_dir
    run(cfg, copy_files=True, log=lambda *_: None)

    wb = load_workbook(output_dir / "migration_tracker.xlsx")
    ws = wb["Tracker"]
    headers = [c.value for c in ws[1]]
    rows = [dict(zip(headers, r)) for r in ws.iter_rows(min_row=2, values_only=True)]

    row = next(r for r in rows if r["file_name"] == "04_contract_globex.pdf")
    assert row["status"] == "retained"
    assert row["is_duplicate"] is False


def test_near_duplicate_scanned_rescan_is_caught(fixture_pdfs, output_dir):
    """#08 is an exact byte-for-byte copy of #07 (see make_test_pdfs.py's
    module docstring for the full cross-platform rationale) - the match is
    guaranteed via the exact_file hash tier, which is evaluated BEFORE any
    OCR/text extraction and therefore deterministic on every platform
    regardless of OCR engine version/output quality. We deliberately do
    NOT assert a specific dup_method here (only status + duplicate_of) -
    the exact tier that catches it is an implementation detail; what
    matters is that the duplicate is correctly detected and the original
    is correctly retained."""
    cfg = Config()
    cfg.paths.source_dir = fixture_pdfs
    cfg.paths.output_dir = output_dir
    run(cfg, copy_files=True, log=lambda *_: None)

    wb = load_workbook(output_dir / "migration_tracker.xlsx")
    ws = wb["Tracker"]
    headers = [c.value for c in ws[1]]
    rows = {r["file_name"]: r for r in
            (dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True))}

    assert rows["08_scanned_contract_delta_rescan.pdf"]["status"] == "removed"
    assert rows["08_scanned_contract_delta_rescan.pdf"]["duplicate_of"] == "07_scanned_contract_delta.pdf"
    assert rows["07_scanned_contract_delta.pdf"]["status"] == "retained"


def test_scanned_pdfs_are_independently_ocr_processed(fixture_pdfs, output_dir):
    """Companion test to the above: confirms that making #08 an exact file
    copy of #07 did NOT short-circuit real OCR processing for either file -
    both are still independently scanned/flagged and counted in ocr_read,
    exercising the genuine OCR code path end-to-end on every platform."""
    cfg = Config()
    cfg.paths.source_dir = fixture_pdfs
    cfg.paths.output_dir = output_dir
    summary = run(cfg, copy_files=True, log=lambda *_: None)

    wb = load_workbook(output_dir / "migration_tracker.xlsx")
    ws = wb["Tracker"]
    headers = [c.value for c in ws[1]]
    rows = {r["file_name"]: r for r in
            (dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True))}

    for fname in ("07_scanned_contract_delta.pdf", "08_scanned_contract_delta_rescan.pdf"):
        assert rows[fname]["is_scanned"] is True
    # At least the retained copy (07) must have gone through OCR successfully.
    assert summary["scanned"] >= 2


def test_blank_and_zero_byte_files_are_never_merged_with_each_other(fixture_pdfs, output_dir):
    cfg = Config()
    cfg.paths.source_dir = fixture_pdfs
    cfg.paths.output_dir = output_dir
    run(cfg, copy_files=True, log=lambda *_: None)

    wb = load_workbook(output_dir / "migration_tracker.xlsx")
    ws = wb["Tracker"]
    headers = [c.value for c in ws[1]]
    rows = {r["file_name"]: r for r in
            (dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True))}

    assert rows["11_corrupt.pdf"]["status"] == "retained"
    assert rows["12_blank_page.pdf"]["status"] == "retained"


def test_missing_source_folder_raises_system_exit(tmp_path):
    cfg = Config()
    cfg.paths.source_dir = tmp_path / "does_not_exist"
    cfg.paths.output_dir = tmp_path / "out"
    with pytest.raises(SystemExit):
        run(cfg, copy_files=True, log=lambda *_: None)


def test_empty_source_folder_raises_system_exit(tmp_path):
    empty = tmp_path / "empty"
    empty.mkdir()
    cfg = Config()
    cfg.paths.source_dir = empty
    cfg.paths.output_dir = tmp_path / "out"
    with pytest.raises(SystemExit):
        run(cfg, copy_files=True, log=lambda *_: None)


def test_ocr_disabled_still_completes_without_crashing(fixture_pdfs, output_dir):
    cfg = Config()
    cfg.paths.source_dir = fixture_pdfs
    cfg.paths.output_dir = output_dir
    cfg.ingestion.enable_ocr = False
    summary = run(cfg, copy_files=True, log=lambda *_: None)
    assert summary["ocr_read"] == 0
    # Even with OCR disabled, #08 must STILL be caught as a duplicate of #07,
    # since the match happens via the exact_file hash tier, which runs
    # regardless of whether OCR is enabled at all.
    wb = load_workbook(summary["tracker"])
    ws = wb["Tracker"]
    headers = [c.value for c in ws[1]]
    rows = {r["file_name"]: r for r in
            (dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True))}
    assert rows["08_scanned_contract_delta_rescan.pdf"]["status"] == "removed"


def test_progress_callback_reports_every_phase_and_reaches_completion(fixture_pdfs, output_dir):
    cfg = Config()
    cfg.paths.source_dir = fixture_pdfs
    cfg.paths.output_dir = output_dir

    seen_phases = set()
    last_by_phase = {}

    def progress(phase, i, total, name):
        seen_phases.add(phase)
        last_by_phase[phase] = (i, total)
        assert i <= total
        assert i >= 0

    run(cfg, copy_files=True, log=lambda *_: None, progress=progress)

    for phase in ("scan", "extract", "enrich", "dedupe", "organise", "tracker", "done"):
        assert phase in seen_phases
    for phase, (i, total) in last_by_phase.items():
        assert i == total


# --------------------------------------------------------------------------- #
# Unit tests: classify.py
# --------------------------------------------------------------------------- #
def test_classify_document_picks_expected_bucket():
    cfg = ClassificationThresholds()
    contract_text = "SERVICES AGREEMENT\nThis agreement is made between A and B.\nIn witness whereof."
    result = classify_document(contract_text, cfg)
    assert result.doc_type == "Contracts"
    assert result.needs_review is False


def test_classify_document_falls_back_to_other_for_unrelated_text():
    cfg = ClassificationThresholds()
    result = classify_document("This is a completely unrelated memo about lunch plans.", cfg)
    assert result.doc_type == "Other"
    assert result.needs_review is True


def test_classify_document_handles_empty_text():
    cfg = ClassificationThresholds()
    result = classify_document("", cfg)
    assert result.doc_type == "Other"
    assert result.needs_review is True


# --------------------------------------------------------------------------- #
# Unit tests: metadata.py
# --------------------------------------------------------------------------- #
def test_extract_metadata_finds_entity_and_date():
    text = "This Agreement is made between Acme Holdings LLC and Ardent Advisory Ltd.\nDated: 12 January 2024"
    meta = extract_metadata(text)
    assert "Acme Holdings LLC" in meta.entity_name or "Ardent Advisory Ltd" in meta.entity_name
    assert meta.contract_date == "2024-01-12"


def test_extract_metadata_handles_empty_text_without_crashing():
    meta = extract_metadata("")
    assert meta.entity_name == ""
    assert meta.contract_date == ""
    d = meta.to_dict()
    assert "entity_name" in d and "contract_date" in d


# --------------------------------------------------------------------------- #
# Unit tests: dedupe.py
# --------------------------------------------------------------------------- #
def _mk_record(idx, file_name, text_norm, **overrides) -> DedupRecord:
    import hashlib
    defaults = dict(
        idx=idx, file_name=file_name, doc_type="Contracts", text_norm=text_norm,
        file_sha256=hashlib.sha256(file_name.encode()).hexdigest(),
        text_sha256=hashlib.sha256(text_norm.encode()).hexdigest(),
        entity_name="Acme Holdings LLC", contract_date="2024-01-12",
        description="", extraction_method="native_fitz", size_bytes=1000,
        text_len=len(text_norm),
    )
    defaults.update(overrides)
    return DedupRecord(**defaults)


def test_deduplicate_leaves_unique_documents_alone():
    rec_a = _mk_record(0, "a.pdf", "completely unique text about apples " * 10)
    rec_b = _mk_record(1, "b.pdf", "totally different text about oranges " * 10)
    deduplicate([rec_a, rec_b], Dedup())
    assert rec_a.status == "retained"
    assert rec_b.status == "retained"


def test_deduplicate_respects_doc_type_blocking_for_near_duplicates():
    text_a = "services agreement " * 40 + "clause set alpha version one"
    text_b = "services agreement " * 40 + "clause set alpha version two slightly different"
    rec_a = _mk_record(0, "a.pdf", text_a, doc_type="Contracts")
    rec_b = _mk_record(1, "b.pdf", text_b, doc_type="Addendums")
    deduplicate([rec_a, rec_b], Dedup(block_by_type=True, near_dup_similarity=0.80))
    assert rec_a.status == "retained"
    assert rec_b.status == "retained"


def test_deduplicate_exact_match_transcends_doc_type_blocking():
    text = "services agreement " * 40 + "identical content in both copies"
    rec_a = _mk_record(0, "a.pdf", text, doc_type="Contracts")
    rec_b = _mk_record(1, "b.pdf", text, doc_type="Addendums")
    deduplicate([rec_a, rec_b], Dedup(block_by_type=True))
    statuses = {rec_a.status, rec_b.status}
    assert statuses == {"retained", "removed"}


def test_deduplicate_never_merges_two_empty_text_records():
    rec_blank = _mk_record(0, "blank.pdf", "", entity_name="", contract_date="")
    rec_corrupt = _mk_record(1, "corrupt.pdf", "", entity_name="", contract_date="")
    deduplicate([rec_blank, rec_corrupt], Dedup())
    assert rec_blank.status == "retained"
    assert rec_corrupt.status == "retained"


def test_deduplicate_handles_empty_list():
    deduplicate([], Dedup())


def test_deduplicate_progress_callback_reports_granular_progress_and_completes():
    records = [
        _mk_record(i, f"f{i}.pdf", f"unique document body number {i} " * 30, entity_name=f"Entity {i}")
        for i in range(60)
    ]
    calls = []
    deduplicate(records, Dedup(min_chars_for_similarity=10),
                progress=lambda done, total: calls.append((done, total)))
    assert calls
    assert len(calls) > 1
    last_done, last_total = calls[-1]
    assert last_done == last_total


def test_deduplicate_result_identical_regardless_of_batch_size_ordering():
    text_a = "SERVICES AGREEMENT this agreement is made between acme holdings llc and " \
             "ardent advisory ltd dated 12 january 2024 in witness whereof the parties " * 3
    text_b = text_a + " minor addendum clause"
    rec_a = _mk_record(0, "a.pdf", text_a)
    rec_b = _mk_record(1, "b.pdf", text_b)
    deduplicate([rec_a, rec_b], Dedup(near_dup_similarity=0.90, min_chars_for_similarity=20))
    statuses = {rec_a.status, rec_b.status}
    assert statuses == {"retained", "removed"}
